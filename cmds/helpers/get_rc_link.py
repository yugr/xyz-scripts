#!python3

import os
import sys
import tempfile
import shutil
import atexit
import getopt
import time

import imp
try:
  imp.find_module('openpyxl')
except ImportError:
  print("""\
This script needs openpyxl module. You can install it via
  > python -m pip install -U pip openpyxl
""")
  sys.exit(1)

import openpyxl

me = os.path.basename(sys.argv[0])
RCs_name = r'\\fileserver\tools\PreRelease\Releases.xlsm'
RCs_elf_name = r'\\fileserver\tools\PreRelease\Releases_elf.xlsm'
toc_row = 2

def warn(msg):
  sys.stderr.write('%s: warning: %s\n' % (me, msg))

def error(msg):
  sys.stderr.write('%s: error: %s\n' % (me, msg))
  sys.exit(1)

def print_help_and_exit():
  print("""\
Usage: python %s [OPTION]... [REL] CORE
Prints link to release REL for core CORE.
If REL is not specified, downloads the latest "Please Align" release.

Options:
  --help, -h       Print help and exit.
  --verbose, -v    Print debug info (for developers).
  --elf            Print for ELF release (default).
  --coff           Print for COFF release.
  --list           Print list of available releases.

Example:
  $ python %s 20180506.8 s2
  \\\\fileserver\\tools\\PreRelease\\rcs\\XYZ-R2x\\V18\\Main\\20180506.8
""" % (me, me))
  sys.exit(1)
  
def analyze_target(target):
  if '-' in target:
    comps = target.split('-')
    if len(comps) != 2:
      error("invalid target syntax: %s" % target)
    core, abi = comps
  else:
    core = target
    abi = None

  if core.lower() in ('r21', 'r22', 's1', 's2'):
    core = 's'

  sheet_name = core
  if core.lower() in ('r1', 'r2'):
    sheet_name = 'r'

  abi = abi.lower()
  if abi not in ('elf', 'coff'):
    error("unknown ABI: %s" % abi)

  return core, abi, sheet_name

def main():
  try:
    opts, args = getopt.getopt(sys.argv[1:], 'hvl', ['help', 'list', 'verbose', 'coff', 'elf'])
  except getopt.GetoptError as err:
    error(str(err))

  verbose = 0
  abi = None
  do_list = False

  for o, a in opts:
    if o in ('-h', '--help'):
      print_help_and_exit()
    elif o in ('-v', '--verbose'):
      verbose += 1
    elif o == '--elf':
      abi = 'elf'
    elif o == '--coff':
      abi = 'coff'
    elif o in ('-l', '--list'):
      do_list = True
    else:
      assert False, "unhandled option"

  if len(args) == 2:
    rc_name, target = args
  elif len(args) == 1:
    rc_name = None
    target = args[0]
  else:
    error("wrong inputs; for more details run with -h")

  core, abi2, core_sheet_name = analyze_target(target)
  if abi2 is None:
    if abi is None:
      error("please specify ABI via --elf, --coff or CORE")
    is_elf = abi == 'elf'
  elif abi is not None and abi != abi2:
    error("ABIs do not match: %s and %s" % (abi, abi2))
  else:
    is_elf = abi2 == 'elf'

  # Read-only mode suppresses hyperlinks but we must be careful to not change original document.
  # So copy and open as writable.
  fd, temp_xls_name = tempfile.mkstemp(suffix='.xlsm', prefix='get_release_DO_NOT_OPEN')
  os.close(fd)
  retry_count = 0
  xls_name = RCs_elf_name if is_elf else RCs_name
  while True:
    try:
      shutil.copy(xls_name, temp_xls_name)
      break
    except FileNotFoundError:
      retry_count += 1
      if retry_count == 5:
        error("failed to read %s" % xls_name)
      retry_period = 15
      warn("%s is inaccessible, retrying in %d seconds..." % (xls_name, retry_period))
      time.sleep(retry_period)
  atexit.register(lambda: os.unlink(temp_xls_name))

  wb = openpyxl.load_workbook(temp_xls_name)

  for sht_name in wb.sheetnames:
    if sht_name.lower() == core_sheet_name.lower():
      sht = wb[sht_name]
      break
  else:
    error("failed to find sheet which matches '%s' in %s" % (core, xls_name))

  # First read TOC
  toc = {}
  for c in range(1, sht.max_column + 1):
    cell = sht.cell(row=toc_row, column=c)
    if cell.value:
      toc[cell.value] = c

  def get_col(name):
    col = toc.get(name, None)
    if col is None:
      error("failed to find '%s' column in TOC in %s" % (name, xls_name))
    return col

  rc_link_col = get_col('RC Link')
  rc_status_col = get_col('RC Status')

  # Collect all releases
  rcs = {}
  last_rc_name = None
  for r in range(toc_row + 1, sht.max_row + 1):
    link_cell = sht.cell(row=r, column=rc_link_col)
    if not link_cell.value:
      warn("empty 'RC Link' column at row %d" % r)
      break
    name = link_cell.value

    status_cell = sht.cell(row=r, column=rc_status_col)
    is_aligned = status_cell.value and status_cell.value.lower().strip().startswith('to align')

    if link_cell.hyperlink is None:
      warn("no link in link column for release '%s' in %s, skipping" % (name, xls_name))
      continue
    rc_link = link_cell.hyperlink.target

    # Strip leading file:///
    rc_link = rc_link.lstrip('file:')
    rc_link = rc_link.lstrip(rc_link[0])

    rcs[name] = is_aligned, rc_link
    if is_aligned:
      last_rc_name = name

  if do_list:
    for rc_name, (is_aligned, rc_link) in rcs.items():
      if is_aligned:
        print('%s %s' % (rc_name, rc_link))
  elif rc_name is None:  # Take latest release?
    if not last_rc_name:
      error("failed to find latest release in %s" % xls_name)
    rc_name = last_rc_name
    _, rc_link = rcs[rc_name]
    print('%s %s' % (rc_name, rc_link))
  else:
    if rc_name not in rcs:
      error("release %s not found in %s" % (rc_name, xls_name))
    _, rc_link = rcs[rc_name]
    sys.stdout.write(rc_link)

if __name__ == '__main__':
    main()
